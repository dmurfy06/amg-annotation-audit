"""Chunk 6: does weighting by viral abundance change the disputed share?

Every figure so far counts CALLS - one annotated gene, one row. That treats a gene on a
virus present at trace level the same as a gene on the most abundant virus in the sample.
The pre-registration commits to reporting both weightings wherever per-sample abundance
exists, because they answer different questions:

  call-weighted      - how much of the RECORD is disputed
  abundance-weighted - how much of the COMMUNITY the disputed record describes

WHAT CAN AND CANNOT BE DONE, stated up front because it constrains everything below:

  soil        per-sample relative abundance, 9 samples (C1-C3 clean, S1-S3 light,
              S4-S6 heavy), in the same supplementary sheet as the calls
  wastewater  RPKM per viral genome (Dataset S6), mapped to the ORFs encoding each vAMG
  ocean       NO PER-SAMPLE ABUNDANCE IS PUBLISHED

The ocean gap matters more than the other two together: ocean_conservative is 88,729 of
93,413 calls and carries the entire strict-rule result.

TWO TRAPS IN THIS DATA, both of which produced wrong numbers on the first attempt:

1. Abundance is published PER VIRUS, not per gene. A virus carrying 27 AMGs repeats its
   abundance on 27 rows. Summing naively over calls double-counts, and it does so
   unevenly - it inflates whichever group happens to sit on gene-dense genomes. That
   artefact alone produced an apparent 0.58x depletion under the strict rule which
   disappears entirely (0.99x) once each virus is counted once.

2. Viral abundance is heavily right-skewed. An aggregate share can therefore differ from
   the call-weighted share purely because a handful of very abundant viruses fall on one
   side, with no systematic difference between the groups at all. The per-family carrier
   medians below are reported precisely so that aggregate ratios cannot be misread as a
   statement about the families.

Three weightings are computed and reported together:
  per-call     - the naive sum, kept ONLY to show the size of the double-counting bias
  fractional   - each virus's abundance split across the AMGs it carries
  binary       - each virus counted once; abundance of viruses carrying >=1 disputed AMG

Inputs:  data/adjudication_families.tsv, data/harmonised_calls.tsv,
         sources/41396_2022_1188_moesm6_esm.xlsx, sources/es2c07800_si_002.xlsx
Output:  results/chunk6_abundance.txt
"""

import collections
import csv
import math
import re
import statistics
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
FAMILIES = ROOT / "data" / "adjudication_families.tsv"
CALLS = ROOT / "data" / "harmonised_calls.tsv"
SOIL = ROOT / "sources" / "41396_2022_1188_moesm6_esm.xlsx"
WASTE = ROOT / "sources" / "es2c07800_si_002.xlsx"
OUT = ROOT / "results" / "chunk6_abundance.txt"

csv.field_size_limit(10 ** 9)
KO_RE = re.compile(r"K\d{5}")

DOES_NOT_COUNT = ["dcm", "glycoside_hydrolase", "xtmA", "xtmB"]
UNRESOLVABLE = ["dsrC_tusE", "folate", "queuosine"]
RULES = [("Strict", set(DOES_NOT_COUNT)),
         ("Maximally strict", set(DOES_NOT_COUNT + UNRESOLVABLE))]

# Soil sheet columns by POSITION: the header row carries three separate columns all
# named "CV", so a name lookup silently returns the wrong one.
SOIL_SAMPLES = [("C1", 10), ("C2", 11), ("C3", 12),
                ("S1", 14), ("S2", 15), ("S3", 16),
                ("S4", 18), ("S5", 19), ("S6", 20)]


def wilson(k: int, n: int, z: float = 1.96) -> tuple[float, float]:
    if n <= 0:
        return (0.0, 0.0)
    p = k / n
    d = 1 + z * z / n
    centre = p + z * z / (2 * n)
    margin = z * math.sqrt(p * (1 - p) / n + z * z / (4 * n * n))
    return (max(0.0, (centre - margin) / d) * 100, min(1.0, (centre + margin) / d) * 100)


def num(v) -> float:
    """Blank means 'not detected in this sample' - a real zero. The supplement publishes
    a complete matrix, so blanks are not missing data."""
    try:
        return float(v) if v not in (None, "") else 0.0
    except (TypeError, ValueError):
        return 0.0


def load_family_map() -> dict[str, str]:
    ko2fam = {}
    with open(FAMILIES, encoding="utf-8", newline="") as fh:
        for r in csv.DictReader(fh, delimiter="\t"):
            for acc in r["accessions"].split(","):
                if acc.strip():
                    ko2fam[acc.strip()] = r["family"]
    return ko2fam


def load_soil(ko2fam):
    """[(family|None, virus_id, {sample: abundance}), ...] for KO-assigned AMG calls."""
    wb = openpyxl.load_workbook(SOIL, read_only=True, data_only=True)
    rows, no_ko = [], 0
    for i, row in enumerate(wb["AMGs"].iter_rows(values_only=True)):
        if i < 2 or str(row[0] or "").upper() != "AMG":
            continue
        m = KO_RE.search(str(row[8] or ""))
        if not m:
            no_ko += 1
            continue
        rows.append((ko2fam.get(m.group(0)), str(row[2] or ""),
                     {n: num(row[j]) for n, j in SOIL_SAMPLES}))
    wb.close()
    return rows, no_ko


def load_wastewater(ko2fam):
    """[(family|None, genome_id), ...] plus genome -> RPKM."""
    wb = openpyxl.load_workbook(WASTE, read_only=True, data_only=True)
    orf2genome, genome_rpkm = {}, {}
    for i, row in enumerate(wb["Dataset S6"].iter_rows(values_only=True)):
        if i < 2 or not row[0]:
            continue
        genome_rpkm[row[0]] = num(row[3])
        for orf in str(row[6] or "").split(","):
            if orf.strip():
                orf2genome[orf.strip()] = row[0]
    wb.close()
    rows, unmatched = [], 0
    with open(CALLS, encoding="utf-8", newline="") as fh:
        for r in csv.DictReader(fh, delimiter="\t"):
            if r["catalogue"] != "wastewater" or not r["ko"]:
                continue
            g = orf2genome.get(r["gene_id"])
            if g is None:
                unmatched += 1
                continue
            rows.append((ko2fam.get(r["ko"]), g))
    return rows, genome_rpkm, unmatched


def three_weightings(items, fams, abundance_of, unit_of):
    """items: iterable of (family, unit). Returns (per_call, fractional, binary) shares."""
    per_unit = collections.Counter(unit_of(it) for it in items)
    pc_t = sum(abundance_of(it) for it in items)
    pc_e = sum(abundance_of(it) for it in items if it[0] in fams)
    fr_t = sum(abundance_of(it) / per_unit[unit_of(it)] for it in items)
    fr_e = sum(abundance_of(it) / per_unit[unit_of(it)] for it in items if it[0] in fams)
    seen = {unit_of(it): abundance_of(it) for it in items}
    hit = {unit_of(it) for it in items if it[0] in fams}
    bi_t = sum(seen.values())
    bi_e = sum(seen[u] for u in hit)
    pct = lambda a, b: 100 * a / b if b else 0.0
    return pct(pc_e, pc_t), pct(fr_e, fr_t), pct(bi_e, bi_t)


def main() -> None:
    ko2fam = load_family_map()
    out: list[str] = []

    def w(line: str = "") -> None:
        out.append(line)

    w("CHUNK 6 - ABUNDANCE-WEIGHTED vs CALL-WEIGHTED DISPUTED SHARE")
    w("=" * 88)
    w("Call-weighted asks how much of the RECORD is disputed.")
    w("Abundance-weighted asks how much of the COMMUNITY that disputed record describes.")
    w()
    w("!! OCEAN CANNOT BE WEIGHTED. The curated ocean catalogue (Microbiome 2024) publishes")
    w("   annotations without a per-sample abundance table - 88,729 of 93,413 calls, and the")
    w("   catalogue carrying the entire strict-rule result. Nothing below extrapolates to it.")
    w()
    w("Abundance is published PER VIRUS in both catalogues that have it, so the naive sum")
    w("over calls double-counts gene-dense genomes. All three weightings are shown; the")
    w("per-call column is included only to expose the size of that bias.")
    w()

    # ------------------------------------------------ soil
    soil, soil_no_ko = load_soil(ko2fam)
    n_vir = len({v for _, v, _ in soil})
    w("=" * 88)
    w("SOIL (ISME J 2022) - relative abundance (%), 9 samples")
    w("=" * 88)
    w("%d KO-assigned AMG calls on %d distinct viruses (%.2f AMGs per virus); %d AMG rows"
      % (len(soil), n_vir, len(soil) / n_vir, soil_no_ko))
    w("carry no KO and are out of scope for a KO-defined adjudication.")
    w()
    for name, fams in RULES:
        k = sum(1 for f, _, _ in soil if f in fams)
        cw = 100 * k / len(soil)
        lo, hi = wilson(k, len(soil))
        w("%s - call-weighted %.2f%% [%.2f-%.2f] (%d/%d calls)"
          % (name, cw, lo, hi, k, len(soil)))
        means = {}
        for label, idx in (("per-call", 0), ("fractional", 1), ("binary", 2)):
            vals = []
            for s, _ in SOIL_SAMPLES:
                items = [(f, v, a[s]) for f, v, a in soil]
                vals.append(three_weightings(items, fams, lambda it: it[2],
                                             lambda it: it[1])[idx])
            means[label] = sum(vals) / len(vals)
            w("    %-11s mean %5.2f%%  range %5.2f-%5.2f  ratio %.2fx"
              % (label, means[label], min(vals), max(vals),
                 means[label] / cw if cw else 0.0))
        w()

    # ------------------------------------------------ wastewater
    waste, rpkm, unmatched = load_wastewater(ko2fam)
    n_gen = len({g for _, g in waste})
    w("=" * 88)
    w("WASTEWATER (ES&T 2023) - RPKM per viral genome")
    w("=" * 88)
    w("%d KO-assigned calls on %d distinct genomes (%.2f vAMGs per genome); unmatched %d"
      % (len(waste), n_gen, len(waste) / n_gen, unmatched))
    w()
    for name, fams in RULES:
        k = sum(1 for f, _ in waste if f in fams)
        cw = 100 * k / len(waste)
        lo, hi = wilson(k, len(waste))
        pc, fr, bi = three_weightings(waste, fams, lambda it: rpkm[it[1]],
                                      lambda it: it[1])
        w("%s - call-weighted %.2f%% [%.2f-%.2f]" % (name, cw, lo, hi))
        for label, v in (("per-call", pc), ("fractional", fr), ("binary", bi)):
            w("    %-11s %5.2f%%   ratio %s"
              % (label, v, ("%.2fx" % (v / cw)) if cw else "n/a"))
        w()

    # ------------------------------------------------ the control that matters
    w("=" * 88)
    w("IS THE AGGREGATE RATIO A PROPERTY OF THE FAMILIES, OR OF THE ABUNDANCE SKEW?")
    w("=" * 88)
    w("An aggregate ratio below 1.0 can mean disputed families sit on rarer viruses - or")
    w("merely that a few very abundant viruses happen to fall outside them. The direct")
    w("test is the median abundance of the viruses actually carrying each family.")
    w()
    vir = {}
    for fam, v, ab in soil:
        mean_ab = sum(ab.values()) / len(ab)
        vir.setdefault(v, [mean_ab, set()])[1].add(fam)
    disputed = set(DOES_NOT_COUNT + UNRESOLVABLE)
    hit = [a for a, fs in vir.values() if fs & disputed]
    oth = [a for a, fs in vir.values() if not (fs & disputed)]
    w("Soil, one value per virus (mean across the 9 samples):")
    w("  carrying a disputed-family AMG   n=%-4d median %.5f   mean %.5f"
      % (len(hit), statistics.median(hit), statistics.mean(hit)))
    w("  all other AMG viruses            n=%-4d median %.5f   mean %.5f"
      % (len(oth), statistics.median(oth), statistics.mean(oth)))
    w("  ratio of medians %.2fx   ratio of means %.2fx"
      % (statistics.median(hit) / statistics.median(oth),
         statistics.mean(hit) / statistics.mean(oth)))
    w()
    w("  mean/median ratio: disputed %.1fx, other %.1fx - both distributions are heavily"
      % (statistics.mean(hit) / statistics.median(hit),
         statistics.mean(oth) / statistics.median(oth)))
    w("  right-skewed, which is what makes the aggregate share unstable.")
    w()
    w("  Per-family median abundance of carrier viruses, against the overall median:")
    for fam in sorted(disputed):
        xs = [a for a, fs in vir.values() if fam in fs]
        if xs:
            w("    %-20s n=%-4d median %.5f   %.2fx"
              % (fam, len(xs), statistics.median(xs),
                 statistics.median(xs) / statistics.median(oth)))
    w()

    # ------------------------------------------------ conclusion
    w("=" * 88)
    w("WHAT THIS ACTUALLY SHOWS")
    w("=" * 88)
    w("1. The strict-rule 'depletion' is an artefact. Naive per-call summing gives 0.58x;")
    w("   counting each virus once gives 0.99-1.01x. There is no effect. Any figure")
    w("   computed by summing per-virus abundance across calls is wrong in this data.")
    w()
    w("2. Under the maximally-strict rule the abundance-weighted share IS lower than the")
    w("   call-weighted share, in both catalogues and under both correct weightings")
    w("   (soil 0.47-0.49x, wastewater 0.36-0.37x).")
    w()
    w("3. But that is NOT because folate- and queuosine-carrying viruses are rare. Their")
    w("   carrier viruses sit at 0.90x and 0.92x of the overall median - which is to say,")
    w("   typical. The aggregate gap is driven by the right skew of the abundance")
    w("   distribution: a small number of very abundant viruses carry no disputed AMG, and")
    w("   with 57 disputed viruses out of 539 that is not a stable property of the record.")
    w()
    w("   The defensible statement is therefore the narrow one: in these two small")
    w("   catalogues, abundance weighting LOWERS the disputed share, but the families")
    w("   driving the dispute are carried by viruses of ordinary abundance. Claiming that")
    w("   disputed genes sit on rare viruses would over-read the data.")
    w()
    w("4. Direction of effect, declared: this cuts AGAINST the project's hypothesis. The")
    w("   abundance-weighted numbers are smaller than the call-weighted ones, so the")
    w("   headline is weaker under the weighting the pre-registration committed to.")

    OUT.write_text("\n".join(out) + "\n", encoding="utf-8")
    print("\n".join(out))


if __name__ == "__main__":
    main()
