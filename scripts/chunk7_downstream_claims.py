"""Chunk 7: does excluding the disputed genes change a claim somebody actually published?

Everything up to here measures the RECORD. This asks whether the record's problems propagate
into conclusions - the difference between "we found a labelling issue" and "the labelling
issue matters".

Three claims are tested, one from each catalogue's own abstract. They were chosen because
each is quantitative, composition-dependent, and recomputable from the published
supplementary data. No claim was selected after seeing its result.

  OCEAN       "we estimate that ~19% of ocean virus populations carry at least one AMG"
              (Microbiome 2024). A virus whose only AMG is in a disputed family stops
              being AMG-carrying, so this number is directly exposed.

  WASTEWATER  "A total of 101 vAMGs ... the most common of which were the queuosine
              biosynthesis genes folE, queD, and queE and the sulfur metabolism gene cysH"
              (ES&T 2023). The named genes ARE a disputed family.

  SOIL        "the diversity and relative abundance of AMGs significantly increased along
              with the severity of pesticide contamination" (ISME J 2022). A comparative
              claim across a Clean -> Light -> Heavy gradient rather than a compositional
              one - included precisely because it should behave differently.

TWO METHOD POINTS THAT CHANGE THE ANSWERS:

  Namespace.  The ocean test uses BOTH KO and Pfam accessions. `dcm` alone carries 5,797
  KO calls and 5,381 Pfam calls; scaling from the KO subset alone would misstate the
  effect on a claim the paper computed over all 86,913 AMGs.

  Grain.      Soil abundance is published per VIRUS, not per gene (see chunk 6). Every
  soil total here is deduplicated to one value per virus per replicate. Summing over rows
  inflates gene-dense genomes and is wrong in this data.

Inputs:  data/family_accessions.tsv     (disputed families, KO + Pfam)
         data/adjudication_families.tsv (35 families, KO)
         data/harmonised_calls.tsv
         sources/41396_2022_1188_moesm6_esm.xlsx, sources/es2c07800_si_002.xlsx
Output:  results/chunk7_downstream_claims.txt
"""

import collections
import csv
import re
import statistics
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "results" / "chunk7_downstream_claims.txt"
csv.field_size_limit(10 ** 9)
KO_RE = re.compile(r"K\d{5}")

# Verdicts from chunk 5. xtmA/xtmB are DOES NOT COUNT but have 0 ocean and 4 soil calls.
DNC = {"dcm", "glycoside_hydrolase", "xtmA", "xtmB"}
UNRES = {"dsrC_tusE", "folate", "queuosine"}
RULES = [("As published (inclusive)", set()),
         ("Strict", DNC),
         ("Maximally strict", DNC | UNRES)]

PUBLISHED_OCEAN_PCT = 19.0            # the figure in the Microbiome 2024 abstract
SOIL_GROUPS = {"Clean": [10, 11, 12], "Light": [14, 15, 16], "Heavy": [18, 19, 20]}


def num(v) -> float:
    try:
        return float(v) if v not in (None, "") else 0.0
    except (TypeError, ValueError):
        return 0.0


def disputed_accessions() -> dict[str, str]:
    """accession -> family, across BOTH namespaces, for the Martin-named families.

    The status column is load-bearing and must be honoured. Two accessions are EXCLUDED
    by the chunk 2 accession review and are NOT part of any family:
      PF13385  a structural fold (Concanavalin A-like lectin), not a hydrolase activity -
               9,237 ocean calls, 99.4% of the glycoside_hydrolase family before review
      K14652   ribBA, GTP cyclohydrolase II - riboflavin, not folate
    Loading them anyway silently reinstates the exact artefact the review removed, and in
    the ocean test it inflates the dropout by ~7,800 viruses.

    AMBIGUOUS accessions (folE, folE2, queD) ARE kept: they sit at the folate/queuosine
    branch point, both of those families are UNRESOLVABLE, and chunk 2's "wide" reading
    counts them. They cannot change the strict result, where neither family is excluded.
    """
    out = {}
    with open(ROOT / "data" / "family_accessions.tsv", encoding="utf-8", newline="") as fh:
        for r in csv.DictReader(fh, delimiter="\t"):
            if r["status"] == "EXCLUDED":
                continue
            out[r["accession"]] = r["family"]
    return out


def ko_family_map() -> dict[str, str]:
    out = {}
    with open(ROOT / "data" / "adjudication_families.tsv", encoding="utf-8", newline="") as fh:
        for r in csv.DictReader(fh, delimiter="\t"):
            for acc in r["accessions"].split(","):
                if acc.strip():
                    out[acc.strip()] = r["family"]
    return out


# --------------------------------------------------------------- ocean
def test_ocean(w, acc2fam):
    w("=" * 88)
    w("OCEAN (Microbiome 2024) - \"~19% of ocean virus populations carry at least one AMG\"")
    w("=" * 88)
    w("A virus counts as AMG-carrying if it has at least one AMG call. Exclude a family and")
    w("any virus whose calls were ALL in excluded families stops qualifying. Both namespaces")
    w("are used, since dcm carries 5,797 KO and 5,381 Pfam calls in this catalogue.")
    w()
    per = collections.defaultdict(lambda: collections.Counter())
    with open(ROOT / "data" / "harmonised_calls.tsv", encoding="utf-8", newline="") as fh:
        for r in csv.DictReader(fh, delimiter="\t"):
            if r["catalogue"] != "ocean_conservative":
                continue
            fam = None
            for a in (r["ko"], r["pfam_id"]):
                if a in acc2fam:
                    fam = acc2fam[a]
            per[r["contig_id"]][fam] += 1
    n_all = len(per)
    w("virus contigs carrying >=1 AMG call, as published: %d" % n_all)
    w()
    w("%-28s %10s %10s %14s" % ("rule", "still AMG", "lost", "claim becomes"))
    w("-" * 66)
    for label, excl in RULES:
        surv = sum(1 for fams in per.values()
                   if sum(n for f, n in fams.items() if f not in excl) > 0)
        w("%-28s %10d %10d %13.1f%%"
          % (label, surv, n_all - surv, PUBLISHED_OCEAN_PCT * surv / n_all))
    w()
    only = collections.Counter()
    for fams in per.values():
        if sum(n for f, n in fams.items() if f not in (DNC | UNRES)) == 0:
            only["+".join(sorted(f for f in fams if f))] += 1
    w("Viruses whose ENTIRE AMG content is disputed (these are the ones that drop out):")
    for k, v in only.most_common(6):
        w("    %-30s %d" % (k or "(none)", v))
    w()
    surv_max = sum(1 for fams in per.values()
                   if sum(n for f, n in fams.items() if f not in (DNC | UNRES)) > 0)
    w("The claim moves from %.0f%% to %.1f%%. Note what carries that: %d viruses qualify as"
      % (PUBLISHED_OCEAN_PCT, PUBLISHED_OCEAN_PCT * surv_max / n_all,
         only.get("dcm", 0)))
    w("AMG-carrying solely on the strength of `dcm` - a family whose verdict the protocol")
    w("fixed in advance as a worked example, and which Martin et al. name explicitly.")
    w()


# --------------------------------------------------------------- wastewater
def test_wastewater(w, ko2fam):
    w("=" * 88)
    w("WASTEWATER (ES&T 2023) - \"the most common ... queuosine genes folE, queD, queE\"")
    w("=" * 88)
    w("Abstract: \"A total of 101 vAMGs involved in various metabolic pathways were")
    w("identified, the most common of which were the queuosine biosynthesis genes folE,")
    w("queD, and queE and the sulfur metabolism gene cysH.\"")
    w()
    wb = openpyxl.load_workbook(ROOT / "sources" / "es2c07800_si_002.xlsx",
                                read_only=True, data_only=True)
    rows = []
    for i, row in enumerate(wb["Dataset S4"].iter_rows(values_only=True)):
        if i < 3 or not row[0]:
            continue
        ko = str(row[1] or "NA")
        rows.append((str(row[9] or "NA").strip(), ko2fam.get(ko)))
    wb.close()
    for label, excl in RULES:
        keep = [r for r in rows if r[1] not in excl]
        counts = collections.Counter(g for g, _ in keep)
        w("%s - %d of %d vAMGs remain" % (label, len(keep), len(rows)))
        for gene, n in counts.most_common(5):
            fam = next((f for g, f in keep if g == gene), None)
            w("    %-14s %3d  (%4.1f%%)   %s" % (gene, n, 100 * n / len(keep), fam or ""))
        w()
    w("The paper's named genes, as published:")
    counts = collections.Counter(g for g, _ in rows)
    queue = sum(n for g, n in counts.items()
                if any(g.startswith(p) for p in ("queD", "queE", "queC", "queF", "GCH1")))
    w("    queuosine pathway genes, aggregated: %d" % queue)
    w("    cysH (not a disputed family):        %d" % counts.get("cysH", 0))
    w()
    w("The claim is only true if the queuosine genes are aggregated as a pathway: no single")
    w("queuosine gene beats cysH's 12, but together they reach %d. Removing the family" % queue)
    w("removes the subject of the sentence entirely, and 28.7% of the catalogue with it.")
    w()


# --------------------------------------------------------------- soil
def test_soil(w, ko2fam):
    w("=" * 88)
    w("SOIL (ISME J 2022) - AMG abundance and diversity rise with contamination severity")
    w("=" * 88)
    w("Abstract: \"the diversity and relative abundance of AMGs significantly increased")
    w("along with the severity of pesticide contamination\". Gradient: Clean -> Light -> Heavy,")
    w("three replicates each. Abundance is per VIRUS, so every total is deduplicated.")
    w()
    wb = openpyxl.load_workbook(ROOT / "sources" / "41396_2022_1188_moesm6_esm.xlsx",
                                read_only=True, data_only=True)
    rows = []
    for i, row in enumerate(wb["AMGs"].iter_rows(values_only=True)):
        if i < 2 or str(row[0] or "").upper() != "AMG":
            continue
        m = KO_RE.search(str(row[8] or ""))
        rows.append((ko2fam.get(m.group(0)) if m else None, str(row[2] or ""),
                     {g: [num(row[j]) for j in idx] for g, idx in SOIL_GROUPS.items()}))
    wb.close()

    for metric in ("abundance", "richness"):
        w("%s:" % metric.upper())
        for label, excl in RULES:
            means = {}
            for g in SOIL_GROUPS:
                reps = []
                for k in range(3):
                    seen = {}
                    for fam, virus, ab in rows:
                        if fam in excl:
                            continue
                        seen[virus] = max(seen.get(virus, 0.0), ab[g][k])
                    reps.append(sum(seen.values()) if metric == "abundance"
                                else sum(1 for x in seen.values() if x > 0))
                means[g] = (statistics.mean(reps), statistics.stdev(reps))
            w("  %-26s Clean %7.2f+-%.2f   Light %7.2f+-%.2f   Heavy %7.2f+-%.2f   H/C %.2fx"
              % (label, means["Clean"][0], means["Clean"][1],
                 means["Light"][0], means["Light"][1],
                 means["Heavy"][0], means["Heavy"][1],
                 means["Heavy"][0] / means["Clean"][0]))
        w()


def main() -> None:
    out: list[str] = []

    def w(line: str = "") -> None:
        out.append(line)

    w("CHUNK 7 - DO THE DISPUTED GENES CHANGE A PUBLISHED CLAIM?")
    w("=" * 88)
    w("One claim per catalogue, taken from its own abstract, each recomputed under the")
    w("adjudication's four rules. Claims were fixed before any was recomputed.")
    w()

    acc2fam = disputed_accessions()
    ko2fam = ko_family_map()
    test_ocean(w, acc2fam)
    test_wastewater(w, ko2fam)
    test_soil(w, ko2fam)

    w("=" * 88)
    w("THE PATTERN, AND IT IS THE RESULT")
    w("=" * 88)
    w("The three claims are affected very differently, and the ordering is not arbitrary -")
    w("it tracks how specifically each claim depends on the disputed families.")
    w()
    w("  DESTROYED - wastewater. The claim names the queuosine genes as the most common")
    w("    vAMGs. Those genes are the disputed family. 29 of 101 vAMGs are removed and the")
    w("    subject of the sentence no longer exists; cysH inherits the top of the list.")
    w("    A claim ABOUT a disputed family cannot survive that family being disputed.")
    w()
    w("  MOVED, MODESTLY - ocean. \"~19% of virus populations carry at least one AMG\"")
    w("    becomes 17.4% (strict) or 16.9% (maximally strict) - a relative fall of about")
    w("    11%. Real, worth reporting, not fatal. 4,945 virus populations are AMG-carrying")
    w("    on the strength of `dcm` alone.")
    w()
    w("  UNTOUCHED - soil. The contamination gradient holds at 1.25x abundance and 2.54x")
    w("    richness under every rule. Disputed families are a roughly constant fraction")
    w("    across the gradient, so they cancel in the comparison.")
    w()
    w("The generalisation worth making: disputed families CANCEL in comparisons and")
    w("PERSIST in descriptions. A paper asking 'does AMG content differ between conditions?'")
    w("is largely safe. A paper asking 'what are these viruses doing?' is answering with")
    w("families the evidence cannot currently resolve - and the more specific the answer,")
    w("the more of it rests on them.")
    w()
    w("Direction of effect, declared: every recomputation here REDUCES the published")
    w("figure. None of the three claims is strengthened by the adjudication. Equally, none")
    w("of the three papers did anything the field would call wrong - they applied the")
    w("standard category. That is the point.")

    OUT.write_text("\n".join(out) + "\n", encoding="utf-8")
    print("\n".join(out))


if __name__ == "__main__":
    main()
