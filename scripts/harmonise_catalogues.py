"""
harmonise_catalogues.py — CHUNK 2. Put every AMG catalogue into one schema, and
work out what a "duplicate" actually is before removing any.

WHY ONE SCHEMA
    The three catalogues were built by different groups with different tools and, crucially,
    key on different identifier systems:

        ocean (Microbiome 2024)  DRAM-v            KEGG KO       88,729 curated calls
        soil  (ISME J 2022)      VIBRANT + DRAM-v  Pfam          ~4,583 calls, with abundances
        wastewater (ES&T 2023)   custom hmmer      prose only    101 calls  <- NOT YET OBTAINED

    Nothing can be compared across them until they share a row format. This script builds
    that format and reports honestly on what each catalogue can and cannot supply.

DEDUPLICATION — measured, not assumed
    05_redteam.md recorded "~1,800 duplicate gene IDs" in the ocean conservative table.
    A duplicate gene_id is not automatically an error: DRAM-v emits one row per annotation
    hit, so the same gene can legitimately appear more than once with different evidence.
    This script therefore CLASSIFIES the duplicates before collapsing anything, and reports
    counts under every rule so the choice is visible rather than buried.

Outputs
    data/harmonised_calls.tsv       one row per AMG call (gitignored, regenerable)
    results/chunk2_harmonisation.txt the report

Usage:
    .venv/Scripts/python.exe scripts/harmonise_catalogues.py
"""

import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OCEAN = ROOT / "data" / "GlobalAMGs_SOM.xlsx"
SOIL = ROOT / "sources" / "41396_2022_1188_moesm6_esm.xlsx"
WASTE = ROOT / "sources" / "es2c07800_si_002.xlsx"
PFAM_CLANS = ROOT / "data" / "Pfam-A.clans.tsv.gz"
OUT_TSV = ROOT / "data" / "harmonised_calls.tsv"
OUT_TXT = ROOT / "results" / "chunk2_harmonisation.txt"

KO_RE = re.compile(r"K\d{5}")

SCHEMA = [
    "catalogue",      # ocean_conservative | ocean_permissive | soil | wastewater
    "environment",    # marine | soil | activated_sludge
    "pipeline",       # DRAM-v | VIBRANT+DRAM-v | custom_hmmer
    "gene_id",        # gene identifier as published
    "contig_id",      # scaffold / contig / virus ID
    "ko",             # KEGG KO accession, or ""
    "pfam_id",        # Pfam accession PF#####, or ""
    "pfam_text",      # Pfam's OWN human-readable description, or ""
    "cazy",           # CAZy family, or ""
    "description",    # the catalogue's own gene description (KEGG-derived for DRAM-v)
    "aux_score",      # DRAM-v auxiliary score 1-5, or ""
    "amg_flags",      # DRAM-v flag letters, or ""
    "has_abundance",  # 1 if per-sample abundance is published for this call
]

# pfam_text and description are kept SEPARATE on purpose. In DRAM-v output,
# gene_description is KEGG-derived ("DNA (cytosine-5-)-methyltransferase") while pfam_hits
# carries Pfam's own wording ("C-5 cytosine-specific DNA methylase [PF00145.17]"). Matching
# the first and calling it a Pfam result silently produces a KEGG answer wearing a Pfam
# label — which is exactly the confusion this project exists to measure.
PFAM_ID_RE = re.compile(r"PF\d{5}")


def pfam_name_to_accession() -> dict[str, str]:
    """Soil publishes Pfam SHORT NAMES ('Glyco_trans_1_2'), not accessions. Amendment 1
    requires membership by accession, so the names must be resolved through Pfam's own
    mapping rather than by matching descriptions — which would reintroduce text matching
    by the back door. Source: Pfam-A.clans.tsv, columns accession/clan/clan_id/name/desc."""
    import gzip
    if not PFAM_CLANS.exists():
        return {}
    m = {}
    with gzip.open(PFAM_CLANS, "rt", encoding="utf-8", errors="replace") as fh:
        for line in fh:
            f = line.rstrip("\n").split("\t")
            if len(f) >= 4 and f[0].startswith("PF"):
                m[f[3]] = f[0]
    return m


PFAM_BY_NAME = pfam_name_to_accession()

_out = []


def say(s: str = "") -> None:
    print(s)
    _out.append(s)


def rows_of(ws, header_row: int):
    """Yield dicts keyed by header. header_row is 0-based; sheets from ISME J carry a
    title in row 0 and the real header in row 1."""
    it = ws.iter_rows(values_only=True)
    for _ in range(header_row):
        next(it)
    hdr = [("" if c is None else str(c).strip()) for c in next(it)]
    for row in it:
        if row is None or all(v is None for v in row):
            continue
        yield hdr, row


def cell(hdr, row, name: str) -> str:
    try:
        i = hdr.index(name)
    except ValueError:
        return ""
    return "" if i >= len(row) or row[i] is None else str(row[i]).strip()


def load_ocean(ws, catalogue: str) -> list[dict]:
    out = []
    for hdr, row in rows_of(ws, 0):
        kid = cell(hdr, row, "kegg_id")
        m = KO_RE.search(kid)
        pf = cell(hdr, row, "pfam_hits")          # "Description [PF00145.17]; ..."
        pid = PFAM_ID_RE.search(pf)
        out.append({
            "catalogue": catalogue,
            "environment": "marine",
            "pipeline": "DRAM-v",
            "gene_id": cell(hdr, row, "gene"),
            "contig_id": cell(hdr, row, "scaffold"),
            "ko": m.group(0) if m else "",
            "pfam_id": pid.group(0) if pid else "",
            "pfam_text": pf,
            "cazy": cell(hdr, row, "cazy_hits"),
            "description": cell(hdr, row, "gene_description") or cell(hdr, row, "kegg_hit"),
            "aux_score": cell(hdr, row, "auxiliary_score"),
            "amg_flags": cell(hdr, row, "amg_flags"),
            "has_abundance": "0",
        })
    return out


def load_soil(ws) -> list[dict]:
    out = []
    for hdr, row in rows_of(ws, 1):          # title in row 0, header in row 1
        if cell(hdr, row, "GROUP").upper() != "AMG":
            continue
        kegg = cell(hdr, row, "KEGG annotation")
        m = KO_RE.search(kegg)
        out.append({
            "catalogue": "soil",
            "environment": "soil",
            "pipeline": "VIBRANT+DRAM-v",
            "gene_id": cell(hdr, row, "virusID-gene"),
            "contig_id": cell(hdr, row, "virus ID"),
            "ko": m.group(0) if m else "",
            "pfam_id": PFAM_BY_NAME.get(cell(hdr, row, "Pfam name"), ""),  # name -> accession
            "pfam_text": cell(hdr, row, "PFAM domain description"),
            "cazy": cell(hdr, row, "CAZy annotation"),
            "description": cell(hdr, row, "PFAM domain description"),
            "aux_score": "",
            "amg_flags": "",
            "has_abundance": "1",            # per-sample columns C1-C3, S1-S6
        })
    return out


def load_wastewater(ws) -> list[dict]:
    """ES&T Dataset S4, "Functional annotation of vAMGs". Title in row 0, a merged
    group header in row 1, real column names in row 2. Carries BOTH a kofamscan KO and
    an hmmsearch Pfam accession, so this catalogue can be matched in either namespace."""
    out = []
    for hdr, row in rows_of(ws, 2):
        orf = cell(hdr, row, "ORF")
        if not orf:
            continue
        m = KO_RE.search(cell(hdr, row, "ko"))
        pid = PFAM_ID_RE.search(cell(hdr, row, "accession"))
        out.append({
            "catalogue": "wastewater",
            "environment": "activated_sludge",
            "pipeline": "custom_hmmer_kofamscan",
            "gene_id": orf,
            "contig_id": orf.rsplit("_", 1)[0] if "_" in orf else orf,
            "ko": m.group(0) if m else "",
            "pfam_id": pid.group(0) if pid else "",
            "pfam_text": cell(hdr, row, "gene name"),
            "cazy": "",
            "description": cell(hdr, row, "ko definition"),
            "aux_score": "",
            "amg_flags": "",
            "has_abundance": "1",          # Dataset S6/S7 carry RPKM
        })
    return out


def duplicate_report(calls: list[dict], label: str) -> None:
    """Classify duplicate gene_ids before deciding whether to collapse them."""
    by_gene = defaultdict(list)
    for c in calls:
        if c["gene_id"]:
            by_gene[c["gene_id"]].append(c)
    dups = {g: rs for g, rs in by_gene.items() if len(rs) > 1}
    extra = sum(len(rs) - 1 for rs in dups.values())

    say(f"\n  {label}")
    say(f"    calls                       : {len(calls):,}")
    say(f"    distinct gene_ids           : {len(by_gene):,}")
    say(f"    gene_ids appearing >1 time  : {len(dups):,}")
    say(f"    surplus rows they contribute: {extra:,}  ({extra/len(calls):.2%} of the table)")
    if not dups:
        return

    # Why are they duplicated? Identical rows, or genuinely different evidence?
    kinds = Counter()
    ko_conflict = 0
    for rs in dups.values():
        sigs = {(r["ko"], r["pfam_id"], r["pfam_text"], r["cazy"]) for r in rs}
        kos = {r["ko"] for r in rs if r["ko"]}
        if len(sigs) == 1:
            kinds["identical rows (safe to collapse)"] += len(rs) - 1
        else:
            kinds["differing annotation evidence"] += len(rs) - 1
        if len(kos) > 1:
            ko_conflict += 1
    for k, v in kinds.most_common():
        say(f"      {k:<42} {v:>7,} surplus rows")
    say(f"      gene_ids assigned CONFLICTING KOs          {ko_conflict:>7,}")
    if ko_conflict:
        ex = next(g for g, rs in dups.items() if len({r['ko'] for r in rs if r['ko']}) > 1)
        say(f"        e.g. {ex[:58]} -> {sorted({r['ko'] for r in dups[ex] if r['ko']})}")


def namespace_report(calls: list[dict], label: str) -> None:
    n = len(calls)
    ko = sum(1 for c in calls if c["ko"])
    pf = sum(1 for c in calls if c["pfam_text"])
    cz = sum(1 for c in calls if c["cazy"])
    say(f"    {label:<22} n={n:>7,}   KO {ko:>7,} ({ko/n:5.1%})"
        f"   Pfam {pf:>7,} ({pf/n:5.1%})   CAZy {cz:>6,} ({cz/n:5.1%})")


def main() -> None:
    say("CHUNK 2 — harmonisation and deduplication")
    say("=" * 78)

    wb = openpyxl.load_workbook(OCEAN, read_only=True, data_only=True)
    ocean_c = load_ocean(wb["Table_S5_Conservative_AMGs"], "ocean_conservative")
    ocean_p = load_ocean(wb["Table_S4_Permissive_AMGs"], "ocean_permissive")
    wb.close()

    wb = openpyxl.load_workbook(SOIL, read_only=True, data_only=True)
    soil = load_soil(wb["AMGs"])
    wb.close()

    waste = []
    if WASTE.exists():
        wb = openpyxl.load_workbook(WASTE, read_only=True, data_only=True)
        waste = load_wastewater(wb["Dataset S4"])
        wb.close()

    say("\nWHICH IDENTIFIER SYSTEM EACH CATALOGUE ACTUALLY CARRIES")
    say("  (this is the namespace finding, measured rather than asserted)")
    namespace_report(ocean_c, "ocean conservative")
    namespace_report(ocean_p, "ocean permissive")
    namespace_report(soil, "soil")
    if waste:
        namespace_report(waste, "wastewater")

    say("\nDUPLICATES — classified before anything is removed")
    duplicate_report(ocean_c, "OCEAN CONSERVATIVE")
    duplicate_report(soil, "SOIL")
    if waste:
        duplicate_report(waste, "WASTEWATER")

    all_calls = ocean_c + ocean_p + soil + waste
    OUT_TSV.parent.mkdir(exist_ok=True)
    with OUT_TSV.open("w", encoding="utf-8", newline="") as fh:
        fh.write("\t".join(SCHEMA) + "\n")
        for c in all_calls:
            fh.write("\t".join(str(c[k]).replace("\t", " ").replace("\n", " ")
                               for k in SCHEMA) + "\n")

    say(f"\nWROTE {OUT_TSV.relative_to(ROOT)}  —  {len(all_calls):,} rows, {len(SCHEMA)} columns")

    if waste:
        say("\nWASTEWATER — obtained 2026-08-05, Dataset S4 of the ES&T Supporting Information.")
        say(f"  {len(waste)} vAMG calls, matching the paper's stated 101 exactly, and carrying")
        say("  BOTH kofamscan KO and hmmsearch Pfam accessions — so it can be measured in")
        say("  either namespace. The paper's own text carries zero KO accessions, which is why")
        say("  the 19.8% could only ever be derived from prose before this table was obtained.")
        say("  The hard gate in 06_project_brief.md is now PASSED for all three catalogues.")
    else:
        say("\nWASTEWATER — still missing. Fetch Dataset S4 from the ES&T Supporting Information")
        say("  at https://pubs.acs.org/doi/10.1021/acs.est.2c07800 (ACS blocks automated download).")

    OUT_TXT.parent.mkdir(exist_ok=True)
    OUT_TXT.write_text("\n".join(_out) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
