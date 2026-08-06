"""
inspect_catalogues.py — CHUNK 2, step 0. What is actually in each catalogue file?

House rule 2: verify the data exists in the shape the question needs, BEFORE designing
anything. So this script designs nothing. It prints sheet names, row counts, headers and
a sample row for every candidate supplementary workbook, so the harmonised schema can be
built from what is really there rather than from what the papers say is there.

The hard gate from 06_project_brief.md:
    "If catalogues 2 and 3 do not publish per-gene tables carrying KO or PFAM identifiers,
     the cross-dataset claim is impossible and the project reduces to a single-catalogue
     report. Test this first."

Usage:
    .venv/Scripts/python.exe scripts/inspect_catalogues.py
"""

from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent

# Every workbook that might carry a per-gene AMG table.
CANDIDATES = [
    (ROOT / "data" / "GlobalAMGs_SOM.xlsx", "OCEAN — Microbiome 2024, Zenodo 12668289"),
    (ROOT / "sources" / "41396_2022_1188_moesm2_esm.xlsx", "SOIL — ISME J 2022 supp 2"),
    (ROOT / "sources" / "41396_2022_1188_moesm3_esm.xlsx", "SOIL — ISME J 2022 supp 3"),
    (ROOT / "sources" / "41396_2022_1188_moesm4_esm.xlsx", "SOIL — ISME J 2022 supp 4"),
    (ROOT / "sources" / "41396_2022_1188_moesm5_esm.xlsx", "SOIL — ISME J 2022 supp 5"),
    (ROOT / "sources" / "41396_2022_1188_moesm6_esm.xlsx", "SOIL — ISME J 2022 supp 6"),
    (ROOT / "sources" / "41396_2022_1188_moesm7_esm.xlsx", "SOIL — ISME J 2022 supp 7"),
]

# Identifier columns that would let a row be classified against the rubric.
ID_HINTS = ("ko", "kegg", "pfam", "cazy", "ec_", "eggnog", "vog", "uniprot")


def describe(path: Path, label: str) -> None:
    print(f"\n{'=' * 78}\n{label}\n  {path.name}  ({path.stat().st_size / 1e6:.1f} MB)\n{'=' * 78}")
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 - want the reason, whatever it is
        print(f"  COULD NOT OPEN: {type(e).__name__}: {e}")
        return

    for name in wb.sheetnames:
        ws = wb[name]
        it = ws.iter_rows(values_only=True)
        try:
            hdr = [("" if c is None else str(c).strip()) for c in next(it)]
        except StopIteration:
            print(f"\n  [{name}]  EMPTY")
            continue

        # Count rows and find the first fully populated one, without loading the sheet.
        n = 0
        sample = None
        for row in it:
            if row is None or all(v is None for v in row):
                continue
            n += 1
            if sample is None:
                sample = row
        print(f"\n  [{name}]  {n:,} data rows, {len(hdr)} columns")

        flagged = [h for h in hdr if any(k in h.lower() for k in ID_HINTS)]
        print(f"    identifier-ish columns: {flagged if flagged else 'NONE — cannot classify rows'}")
        print(f"    all columns: {hdr[:14]}{' …' if len(hdr) > 14 else ''}")
        if sample:
            pairs = [f"{h}={str(v)[:34]}" for h, v in zip(hdr, sample) if v is not None][:7]
            print(f"    first row: {pairs}")
    wb.close()


def main() -> None:
    print("CHUNK 2 step 0 — inventory only. Designs nothing, changes nothing.")
    for path, label in CANDIDATES:
        if path.exists():
            describe(path, label)
        else:
            print(f"\n{'=' * 78}\n{label}\n  MISSING: {path}\n{'=' * 78}")

    print(f"\n{'=' * 78}")
    print("WASTEWATER — ES&T 2023")
    print(f"{'=' * 78}")
    pdf = ROOT / "sources" / "es2c07800.pdf"
    txt = ROOT / "refs" / "est_2023_wastewater_amg.txt"
    print(f"  paper PDF : {'present' if pdf.exists() else 'MISSING'}")
    print(f"  paper text: {'present' if txt.exists() else 'MISSING'}")
    print("  supplementary workbook: NOT PRESENT in sources/")
    print("  -> The 19.8% (20/101) figure came from the paper's prose, not a per-gene table.")
    print("     Whether ES&T published a per-gene SI table is the open question for this chunk.")


if __name__ == "__main__":
    main()
